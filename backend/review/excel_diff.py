#review/excel_diff.py
import pandas as pd
import json
from openpyxl import load_workbook
from memory.corrections import record_human_correction

LEARNABLE_COLUMNS = ["Name", "Service", "Amount"]

def diff_and_learn(corrected_path):
    """
    Opens the corrected Excel, extracts original AI data from the hidden 
    metadata sheet, and learns the differences.
    """
    print(f"🧐 Analyzing corrections in: {corrected_path.name}")
    
    # 1. Load the Excel using openpyxl to get the hidden metadata
    wb = load_workbook(corrected_path, data_only=True)
    
    if "METADATA" not in wb.sheetnames:
        print("❌ Error: No METADATA sheet found. Cannot learn from this file.")
        return

    ws_meta = wb["METADATA"]
    tenant_id = ws_meta.cell(row=1, column=2).value
    original_data_json = ws_meta.cell(row=2, column=2).value
    
    # 2. Convert metadata back to a list of dictionaries
    original_rows = json.loads(original_data_json)

    # 3. Read the 'Invoice Audit' sheet (where the human made changes)
    # we use header=6 because your excel_writer has several header/summary lines
    # We'll search for the row where 'Name' appears to be safe.
    df_corrected = pd.read_excel(corrected_path, sheet_name="Invoice Audit")
    
    # Find the header row dynamically (looking for "Name" column)
    header_row_index = None
    for i, row in df_corrected.iterrows():
        if "Name" in row.values:
            header_row_index = i
            break
            
    if header_row_index is None:
        print("❌ Error: Could not find data table in Excel.")
        return

    # Re-read with the correct header
    df_corrected = pd.read_excel(corrected_path, sheet_name="Invoice Audit", skiprows=header_row_index + 1)
    df_corrected = df_corrected.fillna("")

    # 4. Compare Original (AI) vs Corrected (Human)
    # We iterate based on the original data length
    learned_count = 0
    for i, orig_row in enumerate(original_rows):
        if i >= len(df_corrected):
            break
            
        corr_row = df_corrected.iloc[i].to_dict()

        # Check each learnable field
        for field in LEARNABLE_COLUMNS:
            orig_val = str(orig_row.get(field, "")).strip()
            # Clean numeric strings to match (e.g., "100.0" vs "100")
            corr_val = str(corr_row.get(field, "")).strip()

            if orig_val != corr_val and corr_val != "":
                # Trigger the versioned learning!
                record_human_correction(
                    original_row=orig_row,
                    corrected_row=corr_row,
                    tenant_id=tenant_id
                )
                learned_count += 1

    if learned_count > 0:
        print(f"✅ Success: Learned {learned_count} new patterns for tenant [{tenant_id}].")
    else:
        print("ℹ️ No changes detected. Nothing new to learn.")