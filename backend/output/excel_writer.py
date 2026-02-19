#output/excel_writer.py
import re
import os
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- STYLING CONSTANTS ---
LOCK_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
OK_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

def write_excel(header_lines, rows, footer, filename, invoice_status=None, review_reasons=None, tenant_id="default_tenant"):
    # ALIGNMENT STEP: Ensure the target directory exists
    Path(filename).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    
    # --- SHEET 1: THE AUDIT (User Facing) ---
    ws = wb.active
    ws.title = "Invoice Audit"

    current_row = 1

    # --- 1. INVOICE HEADER ---
    for line in header_lines:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1, value=line)
        cell.font = Font(bold=True, size=12)
        current_row += 1

    current_row += 1 

    # --- 2. AUDIT SUMMARY ---
    if invoice_status:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1, value=f"SYSTEM STATUS: {invoice_status}")
        cell.font = Font(bold=True)
        cell.fill = OK_FILL if invoice_status == "OK" else LOCK_FILL
        current_row += 1

        if review_reasons:
            for reason in review_reasons:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                ws.cell(row=current_row, column=1, value=f"• {reason}")
                current_row += 1
        current_row += 1

    # --- 3. TABLE HEADERS ---
    columns = ["Name", "Telephone", "Service", "Amount", "Sign", "Review_Status"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    current_row += 1

    # --- 4. TABLE ROWS ---
    for row in rows:
        status = row.get("Review_Status", "OK")
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            
            # Smart Number Handling for Amount Column
            if col_name == "Amount":
                try:
                    # Remove currency symbols/commas to store as float
                    clean_val = float(str(val).replace(",", "").replace("$", ""))
                    cell = ws.cell(row=current_row, column=col_idx, value=clean_val)
                    cell.number_format = '#,##0.00'
                except:
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
            else:
                cell = ws.cell(row=current_row, column=col_idx, value=val)
            
            # Apply Colors
            if status == "CHECK_AMOUNT":
                cell.fill = LOCK_FILL
            elif status == "CHECK_OCR":
                cell.fill = ORANGE_FILL
            elif status == "CHECK_TOTAL":
                cell.fill = YELLOW_FILL
        current_row += 1

    # --- 5. TOTALS SECTION ---
    current_row += 1
    ws.cell(row=current_row, column=3, value="CALCULATED TOTAL:").font = Font(bold=True)
    
    # Use Excel Formula for Totaling (Better than manual sum)
    total_cell = ws.cell(row=current_row, column=4)
    start_data_row = current_row - len(rows)
    end_data_row = current_row - 1
    total_cell.value = f"=SUM(D{start_data_row}:D{end_data_row})"
    total_cell.font = Font(bold=True)
    total_cell.fill = OK_FILL
    total_cell.number_format = '#,##0.00'
    
    current_row += 2

    # --- 6. FOOTER ---
    for key, value in footer.items():
        ws.cell(row=current_row, column=1, value=f"{key}:")
        ws.cell(row=current_row, column=2, value=str(value))
        current_row += 1

    # --- 7. COLUMN WIDTHS ---
    widths = [25, 18, 35, 15, 10, 20]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = width

    # --- 8. NEW FEATURE: HIDDEN METADATA SHEET ---
    # We save a copy of the original data and tenant_id here
    # This makes learning via excel_diff.py much more accurate
    ws_meta = wb.create_sheet("METADATA")
    ws_meta.sheet_state = 'hidden'
    ws_meta.cell(row=1, column=1, value="tenant_id")
    ws_meta.cell(row=1, column=2, value=tenant_id)
    ws_meta.cell(row=2, column=1, value="original_json")
    ws_meta.cell(row=2, column=2, value=json.dumps(rows))

    wb.save(filename)